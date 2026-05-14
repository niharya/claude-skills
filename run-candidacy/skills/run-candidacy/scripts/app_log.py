#!/usr/bin/env python3
"""
app_log.py

Manages the application log. Two sheets:

  Applications: slug | company | role | jd_url | application_url | date_applied
  Outreach:     slug | name | role | linkedin_url | status | last_contacted | notes

Default location is `<skill>/applications.xlsx` (next to the SKILL.md). Override
with the APP_LOG_PATH env var to point it anywhere else (e.g. `~/Documents/applications.xlsx`).

If the target path isn't writable (sandboxed run, permission issue, Excel
lock), entries fall back to `<skill>/pending-log/*.json` so the entry isn't
lost. Merge them back into the xlsx with `app_log.py merge-pending` once the
lock clears.

Usage:
    python3 app_log.py init
    python3 app_log.py applied <slug> <company> <role> <jd_url> [application_url]
    python3 app_log.py outreach <slug> "<name>" "<role>" <linkedin_url>
    python3 app_log.py outreach-from-json <slug> <contacts.json>
    python3 app_log.py status <slug>
    python3 app_log.py list
    python3 app_log.py merge-pending

The xlsx is created on first 'applied' or explicit 'init'. Adding rows is
idempotent — re-adding the same (slug, name) for outreach updates the row.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import sys
from pathlib import Path


SKILL_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_LOG_PATH = SKILL_ROOT / "applications.xlsx"
PENDING_DIR = SKILL_ROOT / "pending-log"


def _resolve_log_path() -> Path:
    """Honor APP_LOG_PATH if set, otherwise use the in-skill default."""
    env = os.environ.get("APP_LOG_PATH", "").strip()
    if env:
        return Path(env).expanduser()
    return DEFAULT_LOG_PATH


LOG_PATH = _resolve_log_path()

APPLICATIONS_HEADERS = [
    "slug", "company", "role", "jd_url", "application_url", "date_applied",
]
OUTREACH_HEADERS = [
    "slug", "name", "role", "linkedin_url", "status", "last_contacted", "notes",
]


def _check_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("app_log.py — openpyxl not installed.")
        print("  pip install --break-system-packages openpyxl")
        sys.exit(0)


def _load_or_create():
    from openpyxl import Workbook, load_workbook

    if LOG_PATH.exists():
        return load_workbook(str(LOG_PATH))

    wb = Workbook()
    apps = wb.active
    apps.title = "Applications"
    apps.append(APPLICATIONS_HEADERS)
    outr = wb.create_sheet("Outreach")
    outr.append(OUTREACH_HEADERS)
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(LOG_PATH))
    return wb


def _save(wb) -> None:
    wb.save(str(LOG_PATH))


# ─── Pending-log fallback ────────────────────────────────────────────────


def _write_pending(kind: str, payload: dict, key: str) -> Path:
    """When the target xlsx isn't writable (sandbox, Excel lock, permissions),
    persist the entry as a JSON file under pending-log/ so it isn't lost."""
    PENDING_DIR.mkdir(parents=True, exist_ok=True)
    safe_key = "".join(c if c.isalnum() or c in "-_." else "_" for c in key)
    path = PENDING_DIR / f"{kind}-{safe_key}.json"
    payload = {**payload, "_kind": kind, "_recorded_at": dt.datetime.now().isoformat(timespec="seconds")}
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    return path


def _is_write_blocked() -> tuple[bool, str]:
    """Quick guess at whether the target log can be written to. Looks for an
    Excel lock file (.~lock.applications.xlsx#) and checks parent writability."""
    lock = LOG_PATH.parent / f".~lock.{LOG_PATH.name}#"
    if lock.exists():
        return True, f"Excel lock file present at {lock}"
    if LOG_PATH.exists() and not os.access(LOG_PATH, os.W_OK):
        return True, f"{LOG_PATH} not writable"
    if not LOG_PATH.exists() and not os.access(LOG_PATH.parent, os.W_OK):
        return True, f"{LOG_PATH.parent} not writable"
    return False, ""


def init() -> None:
    _check_openpyxl()
    _load_or_create()
    print(f"applications.xlsx ready at {LOG_PATH}")


def applied(slug: str, company: str, role: str, jd_url: str, application_url: str = "") -> None:
    payload = {
        "slug": slug, "company": company, "role": role,
        "jd_url": jd_url, "application_url": application_url,
        "date_applied": dt.date.today().isoformat(),
    }
    blocked, reason = _is_write_blocked()
    if blocked:
        path = _write_pending("applied", payload, slug)
        print(f"app_log.py — log not writable ({reason}). Wrote pending entry → {path}")
        print(f"  Merge later with: python3 scripts/app_log.py merge-pending")
        return

    _check_openpyxl()
    try:
        wb = _load_or_create()
        sheet = wb["Applications"]
        for row in sheet.iter_rows(min_row=2, values_only=False):
            if row[0].value == slug:
                row[1].value = company
                row[2].value = role
                row[3].value = jd_url
                row[4].value = application_url
                row[5].value = dt.date.today().isoformat()
                _save(wb)
                print(f"updated row for {slug}")
                return
        sheet.append([slug, company, role, jd_url, application_url, dt.date.today().isoformat()])
        _save(wb)
        print(f"logged {slug} → {LOG_PATH}")
    except (PermissionError, OSError) as e:
        path = _write_pending("applied", payload, slug)
        print(f"app_log.py — write failed ({e}). Wrote pending entry → {path}")
        print(f"  Merge later with: python3 scripts/app_log.py merge-pending")


def outreach(slug: str, name: str, role: str, linkedin_url: str,
             status: str = "pending", last_contacted: str = "", notes: str = "") -> None:
    payload = {
        "slug": slug, "name": name, "role": role,
        "linkedin_url": linkedin_url, "status": status,
        "last_contacted": last_contacted, "notes": notes,
    }
    key = f"{slug}-{name}"
    blocked, reason = _is_write_blocked()
    if blocked:
        path = _write_pending("outreach", payload, key)
        print(f"app_log.py — log not writable ({reason}). Wrote pending entry → {path}")
        return

    _check_openpyxl()
    try:
        wb = _load_or_create()
        sheet = wb["Outreach"]
        for row in sheet.iter_rows(min_row=2, values_only=False):
            if row[0].value == slug and row[1].value == name:
                row[2].value = role
                row[3].value = linkedin_url
                row[4].value = status
                row[5].value = last_contacted
                row[6].value = notes
                _save(wb)
                print(f"updated outreach row for {slug} · {name}")
                return
        sheet.append([slug, name, role, linkedin_url, status, last_contacted, notes])
        _save(wb)
        print(f"added outreach row for {slug} · {name}")
    except (PermissionError, OSError) as e:
        path = _write_pending("outreach", payload, key)
        print(f"app_log.py — write failed ({e}). Wrote pending entry → {path}")


def outreach_from_json(slug: str, contacts_path: Path) -> None:
    contacts = json.loads(contacts_path.read_text(encoding="utf-8"))
    for c in contacts:
        outreach(
            slug,
            c.get("name", ""),
            c.get("role", ""),
            c.get("linkedin_url", ""),
            status=c.get("status", "pending"),
            last_contacted=c.get("last_contacted", ""),
            notes=c.get("notes", ""),
        )


def status(slug: str) -> None:
    _check_openpyxl()
    if not LOG_PATH.exists():
        print(f"{LOG_PATH} doesn't exist yet. Nothing logged.")
        return
    from openpyxl import load_workbook
    wb = load_workbook(str(LOG_PATH))
    apps = wb["Applications"]
    outr = wb["Outreach"]

    found_app = None
    for row in apps.iter_rows(min_row=2, values_only=True):
        if row[0] == slug:
            found_app = row
            break
    if found_app:
        print(f"Application: {found_app[1]} · {found_app[2]} · applied {found_app[5]}")
        print(f"  JD: {found_app[3]}")
        print(f"  Form: {found_app[4]}")
    else:
        print(f"No application logged for slug {slug!r}.")

    outreach_rows = [row for row in outr.iter_rows(min_row=2, values_only=True) if row[0] == slug]
    if outreach_rows:
        print(f"Outreach ({len(outreach_rows)}):")
        for r in outreach_rows:
            print(f"  · {r[1]} — {r[2]} — {r[4]}{(' (' + r[5] + ')') if r[5] else ''}")
    else:
        print("No outreach rows for this slug.")


def merge_pending() -> None:
    if not PENDING_DIR.exists():
        print("No pending-log/ directory. Nothing to merge.")
        return

    pending = sorted(PENDING_DIR.glob("*.json"))
    if not pending:
        print("No pending entries.")
        return

    blocked, reason = _is_write_blocked()
    if blocked:
        print(f"app_log.py — still cannot write to {LOG_PATH} ({reason}). Try again later.")
        return

    _check_openpyxl()
    merged_dir = PENDING_DIR / "merged"
    merged_dir.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%dT%H%M%S")

    count = 0
    for path in pending:
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError) as e:
            print(f"  skip {path.name}: {e}")
            continue
        kind = payload.get("_kind")
        if kind == "applied":
            applied(payload["slug"], payload["company"], payload["role"],
                    payload["jd_url"], payload.get("application_url", ""))
        elif kind == "outreach":
            outreach(payload["slug"], payload["name"], payload["role"],
                     payload["linkedin_url"], status=payload.get("status", "pending"),
                     last_contacted=payload.get("last_contacted", ""),
                     notes=payload.get("notes", ""))
        else:
            print(f"  skip {path.name}: unknown _kind {kind!r}")
            continue
        path.rename(merged_dir / f"{path.stem}.{stamp}.json")
        count += 1
    print(f"merged {count} pending entr{'y' if count == 1 else 'ies'} → {LOG_PATH}")


def list_all() -> None:
    _check_openpyxl()
    if not LOG_PATH.exists():
        print(f"{LOG_PATH} doesn't exist yet.")
        return
    from openpyxl import load_workbook
    wb = load_workbook(str(LOG_PATH))
    apps = wb["Applications"]
    rows = list(apps.iter_rows(min_row=2, values_only=True))
    if not rows:
        print("No applications yet.")
        return
    rows.sort(key=lambda r: r[5] or "", reverse=True)
    for r in rows:
        print(f"  {r[5] or '—'}  ·  {r[0]}  ·  {r[1]} — {r[2]}")


def main() -> int:
    ap = argparse.ArgumentParser(description="Application log management.")
    sub = ap.add_subparsers(dest="cmd", required=True)

    sub.add_parser("init", help="Create applications.xlsx if it doesn't exist.")

    p_applied = sub.add_parser("applied", help="Log an application as submitted.")
    p_applied.add_argument("slug")
    p_applied.add_argument("company")
    p_applied.add_argument("role")
    p_applied.add_argument("jd_url")
    p_applied.add_argument("application_url", nargs="?", default="")

    p_out = sub.add_parser("outreach", help="Add or update one outreach contact.")
    p_out.add_argument("slug")
    p_out.add_argument("name")
    p_out.add_argument("role")
    p_out.add_argument("linkedin_url")
    p_out.add_argument("--status", default="pending")
    p_out.add_argument("--last-contacted", default="", dest="last_contacted")
    p_out.add_argument("--notes", default="")

    p_oj = sub.add_parser("outreach-from-json", help="Bulk-import contacts from JSON.")
    p_oj.add_argument("slug")
    p_oj.add_argument("contacts_json")

    p_status = sub.add_parser("status", help="Show status for one slug.")
    p_status.add_argument("slug")

    sub.add_parser("list", help="List all applications.")
    sub.add_parser("merge-pending", help="Merge pending-log/*.json entries into the xlsx.")

    args = ap.parse_args()

    if args.cmd == "init":
        init()
    elif args.cmd == "applied":
        applied(args.slug, args.company, args.role, args.jd_url, args.application_url)
    elif args.cmd == "outreach":
        outreach(args.slug, args.name, args.role, args.linkedin_url,
                 status=args.status, last_contacted=args.last_contacted, notes=args.notes)
    elif args.cmd == "outreach-from-json":
        outreach_from_json(args.slug, Path(args.contacts_json))
    elif args.cmd == "status":
        status(args.slug)
    elif args.cmd == "list":
        list_all()
    elif args.cmd == "merge-pending":
        merge_pending()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
